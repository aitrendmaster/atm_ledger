"""가계부 내역 Excel 익스포트.

- 월별: 한 시트, 해당 월의 entries + planned + reflections + 합계.
- 연간: 12개 월 시트 + 연간 요약 시트.
"""
from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models.entry import Entry
from ..models.planned import Planned
from ..models.reflection import Reflection

HEADER_FILL = PatternFill("solid", fgColor="FAF7F0")
HEADER_FONT = Font(bold=True, color="2C2418")


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 8
        for cell in col_cells:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(40, len(v) + 2))
        ws.column_dimensions[col_letter].width = max_len


def _write_entries_sheet(ws, entries: list[Entry]) -> int:
    """entries 를 시트에 채우고 총 지출 합계 반환."""
    headers = ["날짜", "내역", "카테고리", "금액(원)", "장소", "주소", "별점", "후기"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    total = 0
    for e in entries:
        ws.append(
            [
                e.date,
                e.description,
                e.category,
                e.amount,
                e.place_name or "",
                e.place_address or "",
                e.rating if e.rating is not None else "",
                e.review or "",
            ]
        )
        total += e.amount or 0

    if entries:
        ws.append([])
        ws.append(["", "", "합계", total, "", "", "", ""])
        ws.cell(row=ws.max_row, column=3).font = HEADER_FONT
        ws.cell(row=ws.max_row, column=4).font = HEADER_FONT

    _autosize(ws)
    return total


def _write_planned_sheet(ws, planned: list[Planned]) -> None:
    headers = ["날짜", "내역", "카테고리", "금액(원)", "유형", "메모"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for p in planned:
        ws.append([p.date, p.description, p.category, p.amount, p.type, p.note or ""])
    _autosize(ws)


def _write_reflections_sheet(ws, reflections: list[Reflection]) -> None:
    headers = ["월", "유형", "내용"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for r in reflections:
        ws.append([r.month, r.type, r.text])
    _autosize(ws)


def build_monthly_xlsx(
    *,
    month: str,
    user_email: str,
    entries: list[Entry],
    planned: list[Planned],
    reflections: list[Reflection],
) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = f"지출 {month}"
    total = _write_entries_sheet(ws, entries)

    if planned:
        _write_planned_sheet(wb.create_sheet(title=f"예정 {month}"), planned)
    if reflections:
        _write_reflections_sheet(wb.create_sheet(title=f"회고 {month}"), reflections)

    info = wb.create_sheet(title="요약", index=0)
    info.append(["Moa AI 가계부 월별 내역"])
    info.append(["계정", user_email])
    info.append(["기간", month])
    info.append(["지출 합계 (원)", total])
    info.append(["지출 건수", len(entries)])
    info.append(["예정 건수", len(planned)])
    info.append(["회고 건수", len(reflections)])
    info.cell(row=1, column=1).font = Font(bold=True, size=14)
    _autosize(info)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_annual_xlsx(
    *,
    year: str,
    user_email: str,
    entries_by_month: dict[str, list[Entry]],
    planned: list[Planned],
    reflections: list[Reflection],
) -> bytes:
    """12개 월 시트 + 요약 시트."""
    wb = Workbook()
    # 첫 자동 시트는 요약으로 재활용
    summary = wb.active
    summary.title = "요약"
    summary.append(["Moa AI 가계부 연간 내역"])
    summary.append(["계정", user_email])
    summary.append(["기간", year])
    summary.append([])
    summary.append(["월", "지출 합계(원)", "지출 건수"])
    for col_idx in range(1, 4):
        c = summary.cell(row=5, column=col_idx)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    year_total = 0
    for month_key in sorted(entries_by_month.keys()):
        rows = entries_by_month[month_key]
        ws = wb.create_sheet(title=f"{month_key}")
        sub_total = _write_entries_sheet(ws, rows)
        year_total += sub_total
        summary.append([month_key, sub_total, len(rows)])

    summary.append([])
    summary.append(["연간 합계 (원)", year_total])
    summary.cell(row=summary.max_row, column=1).font = HEADER_FONT
    summary.cell(row=summary.max_row, column=2).font = HEADER_FONT

    if planned:
        _write_planned_sheet(wb.create_sheet(title="예정"), planned)
    if reflections:
        _write_reflections_sheet(wb.create_sheet(title="회고"), reflections)

    _autosize(summary)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
